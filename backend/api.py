"""
FastAPI 后端 - 提供票据数据接口和静态文件服务

启动方式:
    cd E:/毕业论文/invoice-ocr-system
    uvicorn backend.api:app --reload --port 8000
"""
import json
import time
import uuid
import shutil
import sqlite3
from pathlib import Path

from fastapi import FastAPI, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from backend.database import get_conn, get_db, init_db, add_audit_log, save_field, update_invoice

MAX_UPLOAD_SIZE = 20 * 1024 * 1024  # 20MB

PROJECT_ROOT = Path(__file__).parent.parent
SROIE_IMG_DIR = PROJECT_ROOT / "data" / "sroie" / "data" / "img"
SROIE_KEY_DIR = PROJECT_ROOT / "data" / "sroie" / "data" / "key"
UPLOAD_DIR = PROJECT_ROOT / "data" / "uploads"
FRONTEND_DIR = PROJECT_ROOT / "frontend"

app = FastAPI(title="票据智能识别系统 - 视觉证据追溯")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# 挂载前端静态文件
app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")


def _extract_sample_id(image_path):
    """从数据库中的 image_path 提取样本 ID（如 '000'）"""
    return Path(image_path).stem


def _load_ground_truth(sample_id):
    """加载 SROIE ground truth"""
    key_path = SROIE_KEY_DIR / f"{sample_id}.json"
    if key_path.exists():
        with open(key_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


@app.get("/")
async def index():
    return FileResponse(str(FRONTEND_DIR / "index.html"))


# ========== 统计概览 ==========

@app.get("/api/stats")
async def get_stats():
    """统计概览：总数、各状态数、异常数"""
    with get_db() as conn:
        total = conn.execute("SELECT COUNT(*) FROM invoices").fetchone()[0]
        approved = conn.execute(
            "SELECT COUNT(*) FROM invoices WHERE status = 'approved'"
        ).fetchone()[0]
        rejected = conn.execute(
            "SELECT COUNT(*) FROM invoices WHERE status = 'rejected'"
        ).fetchone()[0]
        pending = conn.execute(
            "SELECT COUNT(*) FROM invoices WHERE status = 'pending'"
        ).fetchone()[0]
        processed = conn.execute(
            "SELECT COUNT(*) FROM invoices WHERE status = 'processed'"
        ).fetchone()[0]
        anomaly = conn.execute(
            "SELECT COUNT(*) FROM invoices WHERE risk_level = 'anomaly' OR risk_flags IS NOT NULL"
        ).fetchone()[0]

        # 最近活动（最新10条审计日志）
        recent_logs = conn.execute(
            "SELECT al.*, i.image_path FROM audit_logs al "
            "LEFT JOIN invoices i ON al.invoice_id = i.id "
            "ORDER BY al.created_at DESC LIMIT 10"
        ).fetchall()
        recent = []
        for r in recent_logs:
            d = dict(r)
            d["sample_id"] = _extract_sample_id(d.get("image_path", "")) if d.get("image_path") else ""
            recent.append(d)

    return {
        "total": total,
        "approved": approved,
        "rejected": rejected,
        "pending": pending,
        "processed": processed,
        "anomaly": anomaly,
        "recent_activity": recent,
    }


# ========== 票据列表（支持 status 筛选）==========

@app.get("/api/invoices")
async def list_invoices(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: str = Query(None),
):
    """分页获取票据列表，支持 status 筛选"""
    with get_db() as conn:
        offset = (page - 1) * page_size

        where = ""
        params = []
        if status == "anomaly":
            where = "WHERE (risk_level = 'anomaly' OR risk_flags IS NOT NULL)"
        elif status:
            where = "WHERE status = ?"
            params.append(status)

        total = conn.execute(
            f"SELECT COUNT(*) FROM invoices {where}", params
        ).fetchone()[0]
        rows = conn.execute(
            f"SELECT id, image_path, status, processing_time_ms, created_at "
            f"FROM invoices {where} ORDER BY created_at DESC LIMIT ? OFFSET ?",
            params + [page_size, offset]
        ).fetchall()

    items = []
    for r in rows:
        sample_id = _extract_sample_id(r["image_path"])
        items.append({
            "id": r["id"],
            "sample_id": sample_id,
            "image_path": r["image_path"],
            "status": r["status"],
            "processing_time_ms": r["processing_time_ms"],
            "created_at": r["created_at"],
        })

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
        "items": items,
    }


# ========== 搜索（必须在 {invoice_id} 之前定义）==========

@app.get("/api/invoices/search")
async def search_invoices(
    q: str = Query("", description="搜索关键词（匹配样本ID或字段值）"),
    status: str = Query(None),
    conf_min: float = Query(None, ge=0, le=1),
    conf_max: float = Query(None, ge=0, le=1),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """高级搜索：关键词 + 状态 + 置信度范围"""
    with get_db() as conn:
        conditions = []
        params = []

        if q:
            conditions.append(
                "(i.image_path LIKE ? OR EXISTS "
                "(SELECT 1 FROM fields f2 WHERE f2.invoice_id = i.id AND f2.final_value LIKE ?))"
            )
            params.extend([f"%{q}%", f"%{q}%"])

        if status == "anomaly":
            conditions.append("(i.risk_level = 'anomaly' OR i.risk_flags IS NOT NULL)")
        elif status:
            conditions.append("i.status = ?")
            params.append(status)

        if conf_min is not None:
            conditions.append(
                "EXISTS (SELECT 1 FROM fields f3 WHERE f3.invoice_id = i.id AND f3.confidence >= ?)"
            )
            params.append(conf_min)

        if conf_max is not None:
            conditions.append(
                "EXISTS (SELECT 1 FROM fields f4 WHERE f4.invoice_id = i.id AND f4.confidence <= ?)"
            )
            params.append(conf_max)

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        offset = (page - 1) * page_size

        total = conn.execute(
            f"SELECT COUNT(*) FROM invoices i {where}", params
        ).fetchone()[0]

        rows = conn.execute(
            f"SELECT i.id, i.image_path, i.status, i.processing_time_ms, i.created_at "
            f"FROM invoices i {where} ORDER BY i.created_at DESC LIMIT ? OFFSET ?",
            params + [page_size, offset]
        ).fetchall()

    items = []
    for r in rows:
        items.append({
            "id": r["id"],
            "sample_id": _extract_sample_id(r["image_path"]),
            "image_path": r["image_path"],
            "status": r["status"],
            "processing_time_ms": r["processing_time_ms"],
            "created_at": r["created_at"],
        })

    return {
        "total": total, "page": page, "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
        "items": items,
    }


# ========== 批量操作（必须在 {invoice_id} 之前定义）==========

class BatchRequest(BaseModel):
    invoice_ids: list[str]
    action: str  # "approve" or "reject"


@app.post("/api/invoices/batch")
async def batch_action(body: BatchRequest):
    """批量审批或拒绝"""
    if body.action not in ("approve", "reject"):
        raise HTTPException(400, "action 必须是 approve 或 reject")
    if not body.invoice_ids:
        raise HTTPException(400, "未选择票据")
    if len(body.invoice_ids) > 500:
        raise HTTPException(400, "单次最多操作 500 条")

    new_status = "approved" if body.action == "approve" else "rejected"
    success_count = 0

    with get_db() as conn:
        for inv_id in body.invoice_ids:
            row = conn.execute("SELECT id, status FROM invoices WHERE id = ?", (inv_id,)).fetchone()
            if not row:
                continue
            conn.execute(
                "UPDATE invoices SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                (new_status, inv_id)
            )
            success_count += 1

    for inv_id in body.invoice_ids:
        add_audit_log(inv_id, f"batch_{body.action}",
                      f"批量{('通过' if body.action == 'approve' else '拒绝')}",
                      actor="reviewer")

    return {"action": body.action, "requested": len(body.invoice_ids),
            "success": success_count, "status": new_status}


# ========== 票据详情 ==========

@app.get("/api/invoices/{invoice_id}")
async def get_invoice_detail(invoice_id: str):
    """获取票据完整详情：字段 + 候选值 + 审计日志"""
    with get_db() as conn:
        # 票据基本信息
        invoice = conn.execute(
            "SELECT * FROM invoices WHERE id = ?", (invoice_id,)
        ).fetchone()
        if not invoice:
            raise HTTPException(404, "Invoice not found")

        invoice_data = dict(invoice)
        sample_id = _extract_sample_id(invoice_data["image_path"])

        # 解析 ocr_raw_json
        ocr_blocks = []
        if invoice_data.get("ocr_raw_json"):
            try:
                ocr_raw = json.loads(invoice_data["ocr_raw_json"])
                ocr_blocks = ocr_raw.get("blocks", [])
            except (json.JSONDecodeError, TypeError):
                pass

        # 字段信息
        field_rows = conn.execute(
            "SELECT * FROM fields WHERE invoice_id = ? ORDER BY id", (invoice_id,)
        ).fetchall()

        fields = []
        for fr in field_rows:
            fd = dict(fr)
            for json_col in ["evidence_bbox", "key_bbox", "rule_details"]:
                if fd.get(json_col):
                    try:
                        fd[json_col] = json.loads(fd[json_col])
                    except (json.JSONDecodeError, TypeError):
                        pass

            cand_rows = conn.execute(
                "SELECT * FROM candidates WHERE field_id = ? ORDER BY final_score DESC",
                (fd["id"],)
            ).fetchall()
            cands = []
            for c in cand_rows:
                cd = dict(c)
                if cd.get("bbox"):
                    try:
                        cd["bbox"] = json.loads(cd["bbox"])
                    except (json.JSONDecodeError, TypeError):
                        pass
                cands.append(cd)
            fd["candidates"] = cands
            fields.append(fd)

        # 审计日志
        log_rows = conn.execute(
            "SELECT * FROM audit_logs WHERE invoice_id = ? ORDER BY created_at",
            (invoice_id,)
        ).fetchall()
        audit_logs = [dict(lr) for lr in log_rows]

    # ground truth
    gt = _load_ground_truth(sample_id)

    return {
        "invoice": {
            "id": invoice_data["id"],
            "sample_id": sample_id,
            "image_path": invoice_data["image_path"],
            "status": invoice_data["status"],
            "created_at": invoice_data["created_at"],
        },
        "ocr_blocks": ocr_blocks,
        "fields": fields,
        "audit_logs": audit_logs,
        "ground_truth": gt,
    }


# ========== 图片服务（同时查找 sroie 和 uploads）==========

@app.get("/api/images/{sample_id}")
async def get_image(sample_id: str):
    """提供票据图片，同时查找 sroie 和 uploads 目录"""
    # 先查 sroie
    img_path = SROIE_IMG_DIR / f"{sample_id}.jpg"
    if img_path.exists():
        return FileResponse(str(img_path), media_type="image/jpeg")

    # 再查 uploads（支持多种扩展名）
    for ext in [".jpg", ".jpeg", ".png", ".bmp", ".tiff"]:
        up_path = UPLOAD_DIR / f"{sample_id}{ext}"
        if up_path.exists():
            media = "image/jpeg" if ext in [".jpg", ".jpeg"] else f"image/{ext[1:]}"
            return FileResponse(str(up_path), media_type=media)

    raise HTTPException(404, f"Image {sample_id} not found")


# ========== 上传处理 ==========

@app.post("/api/upload")
async def upload_invoice(file: UploadFile = File(...)):
    """接收图片 → OCR → 提取 → 存库 → 返回结果"""
    # 确保上传目录存在
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    # 读取并验证文件大小
    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(413, f"文件过大，限制 {MAX_UPLOAD_SIZE // 1024 // 1024}MB")
    if len(content) == 0:
        raise HTTPException(400, "文件为空")

    # 生成唯一文件名
    file_ext = Path(file.filename).suffix.lower() or ".jpg"
    file_id = f"upload_{uuid.uuid4().hex[:12]}"
    save_path = UPLOAD_DIR / f"{file_id}{file_ext}"

    # 保存文件
    with open(save_path, "wb") as f:
        f.write(content)

    start_time = time.time()

    try:
        # OCR
        from backend.ocr_engine import run_ocr
        ocr_result = run_ocr(str(save_path))

        # 字段提取
        from backend.extractor import extract_all_fields
        extraction = extract_all_fields(ocr_result["blocks"], config="full")

        elapsed_ms = (time.time() - start_time) * 1000

        # 存入数据库
        invoice_id = str(uuid.uuid4())
        with get_db() as conn:
            conn.execute(
                "INSERT INTO invoices (id, image_path, ocr_raw_json, status, "
                "processing_time_ms, risk_level) VALUES (?, ?, ?, ?, ?, ?)",
                (invoice_id, f"{file_id}{file_ext}",
                 json.dumps(ocr_result), "processed", elapsed_ms,
                 "anomaly" if extraction["anomalies"] else "normal")
            )

            if extraction["anomalies"]:
                conn.execute(
                    "UPDATE invoices SET risk_flags = ? WHERE id = ?",
                    (json.dumps(extraction["anomalies"]), invoice_id)
                )

        # 保存字段和候选值
        fields_result = extraction["fields"]
        field_name_map = {
            "company": "Company (公司名称)",
            "date": "Date (日期)",
            "address": "Address (地址)",
            "total": "Total (总金额)",
        }

        saved_fields = {}
        for field_key, field_data in fields_result.items():
            candidates_for_db = []
            for c in field_data.get("candidates", []):
                candidates_for_db.append({
                    "source": c.get("source", ""),
                    "value": c.get("value", ""),
                    "ocr_confidence": c.get("ocr_confidence"),
                    "format_score": c.get("format_score"),
                    "cross_field_score": c.get("cross_field_score"),
                    "final_score": c.get("final_score"),
                    "is_selected": c.get("value") == field_data["value"],
                    "bbox": c.get("bbox"),
                })

            fid = save_field(
                invoice_id, field_key,
                field_name_map.get(field_key, field_key),
                field_data["value"],
                field_data["confidence"],
                evidence_bbox=field_data.get("evidence_bbox"),
                decision_reason=field_data.get("decision_reason"),
                candidates_list=candidates_for_db,
            )
            saved_fields[field_key] = {
                "value": field_data["value"],
                "confidence": field_data["confidence"],
            }

        # 审计日志
        add_audit_log(
            invoice_id, "upload_processed", f"上传文件 {file.filename} 处理完成",
            actor="system"
        )

        return {
            "invoice_id": invoice_id,
            "sample_id": file_id,
            "filename": file.filename,
            "processing_time_ms": round(elapsed_ms, 1),
            "fields": saved_fields,
            "anomalies": extraction["anomalies"],
            "status": "processed",
        }

    except Exception as e:
        # 清理失败的上传
        if save_path.exists():
            save_path.unlink()
        raise HTTPException(500, f"处理失败: {str(e)}")


# ========== 审核：通过 ==========

@app.post("/api/invoices/{invoice_id}/approve")
async def approve_invoice(invoice_id: str):
    """审核通过，更新状态 + 审计日志"""
    with get_db() as conn:
        invoice = conn.execute(
            "SELECT id, status FROM invoices WHERE id = ?", (invoice_id,)
        ).fetchone()
        if not invoice:
            raise HTTPException(404, "Invoice not found")

        old_status = invoice["status"]
        conn.execute(
            "UPDATE invoices SET status = 'approved', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (invoice_id,)
        )

    add_audit_log(
        invoice_id, "approve",
        f"票据审核通过 (原状态: {old_status})",
        actor="reviewer"
    )

    return {"status": "approved", "invoice_id": invoice_id}


# ========== 审核：拒绝 ==========

@app.post("/api/invoices/{invoice_id}/reject")
async def reject_invoice(invoice_id: str):
    """标记异常，更新状态 + 审计日志"""
    with get_db() as conn:
        invoice = conn.execute(
            "SELECT id, status FROM invoices WHERE id = ?", (invoice_id,)
        ).fetchone()
        if not invoice:
            raise HTTPException(404, "Invoice not found")

        old_status = invoice["status"]
        conn.execute(
            "UPDATE invoices SET status = 'rejected', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (invoice_id,)
        )

    add_audit_log(
        invoice_id, "reject",
        f"票据审核拒绝 (原状态: {old_status})",
        actor="reviewer"
    )

    return {"status": "rejected", "invoice_id": invoice_id}


# ========== 字段编辑 ==========

class FieldUpdateRequest(BaseModel):
    value: str


@app.put("/api/invoices/{invoice_id}/fields/{field_key}")
async def update_field(invoice_id: str, field_key: str, body: FieldUpdateRequest):
    """人工修正字段值"""
    with get_db() as conn:
        # 验证票据存在
        invoice = conn.execute(
            "SELECT id FROM invoices WHERE id = ?", (invoice_id,)
        ).fetchone()
        if not invoice:
            raise HTTPException(404, "Invoice not found")

        # 查找字段
        field = conn.execute(
            "SELECT id, final_value FROM fields WHERE invoice_id = ? AND field_key = ?",
            (invoice_id, field_key)
        ).fetchone()
        if not field:
            raise HTTPException(404, f"Field '{field_key}' not found")

        old_value = field["final_value"]
        new_value = body.value

        # 更新字段值
        conn.execute(
            "UPDATE fields SET final_value = ? WHERE id = ?",
            (new_value, field["id"])
        )

    # 审计日志
    add_audit_log(
        invoice_id, "field_edit",
        f"字段 {field_key} 从 '{old_value}' 修改为 '{new_value}'",
        actor="reviewer",
        target_field=field_key,
        old_value=old_value,
        new_value=new_value,
    )

    return {
        "invoice_id": invoice_id,
        "field_key": field_key,
        "old_value": old_value,
        "new_value": new_value,
    }


# ========== 分析数据 ==========

@app.get("/api/analytics")
async def get_analytics():
    """分析数据：各字段准确率、置信度分布、处理时间统计、GT对比"""
    with get_db() as conn:
        # 1) 各字段平均置信度
        field_conf = conn.execute(
            "SELECT field_key, AVG(confidence) as avg_conf, "
            "COUNT(*) as cnt, MIN(confidence) as min_conf, MAX(confidence) as max_conf "
            "FROM fields GROUP BY field_key"
        ).fetchall()
        field_confidence = {r["field_key"]: {
            "avg": round(r["avg_conf"] or 0, 4),
            "count": r["cnt"],
            "min": round(r["min_conf"] or 0, 4),
            "max": round(r["max_conf"] or 0, 4),
        } for r in field_conf}

        # 2) 置信度分布（按0.1区间分桶）
        conf_rows = conn.execute("SELECT confidence FROM fields WHERE confidence IS NOT NULL").fetchall()
        conf_buckets = [0] * 10  # [0-0.1), [0.1-0.2), ... [0.9-1.0]
        for r in conf_rows:
            idx = min(int(r["confidence"] * 10), 9)
            conf_buckets[idx] += 1

        # 3) 处理时间统计
        time_stats = conn.execute(
            "SELECT AVG(processing_time_ms) as avg_time, "
            "MIN(processing_time_ms) as min_time, MAX(processing_time_ms) as max_time, "
            "COUNT(*) as cnt FROM invoices WHERE processing_time_ms IS NOT NULL"
        ).fetchone()

        # 处理时间分布
        time_rows = conn.execute(
            "SELECT processing_time_ms FROM invoices WHERE processing_time_ms IS NOT NULL "
            "ORDER BY processing_time_ms"
        ).fetchall()
        time_buckets_labels = ["<200ms", "200-500ms", "500-1s", "1-2s", "2-5s", ">5s"]
        time_buckets = [0] * 6
        for r in time_rows:
            t = r["processing_time_ms"]
            if t < 200: time_buckets[0] += 1
            elif t < 500: time_buckets[1] += 1
            elif t < 1000: time_buckets[2] += 1
            elif t < 2000: time_buckets[3] += 1
            elif t < 5000: time_buckets[4] += 1
            else: time_buckets[5] += 1

        # 4) GT 对比准确率（仅 SROIE 有 GT 的样本）
        all_invoices = conn.execute(
            "SELECT id, image_path FROM invoices"
        ).fetchall()

        gt_results = {"company": {"correct": 0, "total": 0},
                      "date": {"correct": 0, "total": 0},
                      "address": {"correct": 0, "total": 0},
                      "total": {"correct": 0, "total": 0}}

        for inv in all_invoices:
            sample_id = _extract_sample_id(inv["image_path"])
            gt = _load_ground_truth(sample_id)
            if not gt:
                continue
            fields = conn.execute(
                "SELECT field_key, final_value FROM fields WHERE invoice_id = ?",
                (inv["id"],)
            ).fetchall()
            for f in fields:
                fk = f["field_key"]
                if fk not in gt_results:
                    continue
                gt_val = gt.get(fk, "")
                pred_val = f["final_value"] or ""
                gt_results[fk]["total"] += 1
                if pred_val and gt_val:
                    if (pred_val.strip().lower() == gt_val.strip().lower() or
                            pred_val.replace(" ", "") == gt_val.replace(" ", "")):
                        gt_results[fk]["correct"] += 1

        accuracy = {}
        for fk, data in gt_results.items():
            accuracy[fk] = {
                "correct": data["correct"],
                "total": data["total"],
                "rate": round(data["correct"] / data["total"], 4) if data["total"] > 0 else 0,
            }

        # 5) 最近处理时间趋势（最近50条）
        trend_rows = conn.execute(
            "SELECT processing_time_ms, created_at FROM invoices "
            "WHERE processing_time_ms IS NOT NULL ORDER BY created_at DESC LIMIT 50"
        ).fetchall()
        time_trend = [{"time_ms": round(r["processing_time_ms"], 1),
                       "created_at": r["created_at"]} for r in reversed(trend_rows)]

        # 6) 总览统计
        total_invoices = conn.execute("SELECT COUNT(*) FROM invoices").fetchone()[0]
        total_fields = conn.execute("SELECT COUNT(*) FROM fields").fetchone()[0]
        total_candidates = conn.execute("SELECT COUNT(*) FROM candidates").fetchone()[0]
        avg_conf_all = conn.execute("SELECT AVG(confidence) FROM fields WHERE confidence IS NOT NULL").fetchone()[0]
        gt_count = sum(1 for inv in all_invoices if _load_ground_truth(_extract_sample_id(inv["image_path"])))

        # 7) 状态分布
        status_dist = conn.execute(
            "SELECT status, COUNT(*) as cnt FROM invoices GROUP BY status"
        ).fetchall()
        status_distribution = {r["status"]: r["cnt"] for r in status_dist}

        # 8) 总体准确率
        total_correct = sum(d["correct"] for d in gt_results.values())
        total_gt = sum(d["total"] for d in gt_results.values())

    return {
        "overview": {
            "total_invoices": total_invoices,
            "total_fields": total_fields,
            "total_candidates": total_candidates,
            "avg_confidence": round((avg_conf_all or 0) * 100, 1),
            "gt_samples": gt_count,
            "overall_accuracy": round(total_correct / total_gt * 100, 1) if total_gt > 0 else 0,
            "avg_time_ms": round(time_stats["avg_time"] or 0, 1),
        },
        "status_distribution": status_distribution,
        "field_confidence": field_confidence,
        "confidence_distribution": {
            "labels": [f"{i/10:.1f}-{(i+1)/10:.1f}" for i in range(10)],
            "values": conf_buckets,
        },
        "processing_time": {
            "avg": round(time_stats["avg_time"] or 0, 1),
            "min": round(time_stats["min_time"] or 0, 1),
            "max": round(time_stats["max_time"] or 0, 1),
            "count": time_stats["cnt"],
            "distribution": {"labels": time_buckets_labels, "values": time_buckets},
        },
        "accuracy": accuracy,
        "time_trend": time_trend,
    }


# ========== 删除票据 ==========

class DeleteRequest(BaseModel):
    invoice_ids: list[str]


@app.post("/api/invoices/delete")
async def delete_invoices(body: DeleteRequest):
    """删除票据及其关联数据"""
    if not body.invoice_ids:
        raise HTTPException(400, "未选择票据")
    if len(body.invoice_ids) > 500:
        raise HTTPException(400, "单次最多删除 500 条")

    deleted = 0
    with get_db() as conn:
        for inv_id in body.invoice_ids:
            row = conn.execute("SELECT id, image_path FROM invoices WHERE id = ?", (inv_id,)).fetchone()
            if not row:
                continue
            # 删候选值
            field_ids = conn.execute("SELECT id FROM fields WHERE invoice_id = ?", (inv_id,)).fetchall()
            for fid in field_ids:
                conn.execute("DELETE FROM candidates WHERE field_id = ?", (fid["id"],))
            # 删字段、审计日志、票据
            conn.execute("DELETE FROM fields WHERE invoice_id = ?", (inv_id,))
            conn.execute("DELETE FROM audit_logs WHERE invoice_id = ?", (inv_id,))
            conn.execute("DELETE FROM invoices WHERE id = ?", (inv_id,))
            deleted += 1

            # 如果是上传的文件，删除物理文件
            img_path = row["image_path"]
            if img_path:
                p = UPLOAD_DIR / Path(img_path).name
                if p.exists():
                    p.unlink()

    return {"deleted": deleted, "requested": len(body.invoice_ids)}


# ========== 导出 Excel ==========

@app.get("/api/export")
async def export_excel(status: str = Query(None)):
    """导出票据数据为 Excel，可按状态筛选"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with get_db() as conn:
        where = ""
        params = []
        if status:
            where = "WHERE i.status = ?"
            params.append(status)

        rows = conn.execute(
            f"""SELECT i.id, i.image_path, i.status, i.risk_level,
                       i.processing_time_ms, i.created_at, i.updated_at
                FROM invoices i {where}
                ORDER BY i.created_at DESC""",
            params
        ).fetchall()

        # 获取所有字段数据
        all_fields = {}
        for r in rows:
            fields = conn.execute(
                "SELECT field_key, final_value, confidence FROM fields WHERE invoice_id = ?",
                (r["id"],)
            ).fetchall()
            all_fields[r["id"]] = {f["field_key"]: f for f in fields}

    wb = Workbook()
    ws = wb.active
    ws.title = "票据数据"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B4B91", end_color="1B4B91", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = ["样本ID", "状态", "风险等级",
               "公司名称", "公司置信度",
               "日期", "日期置信度",
               "地址", "地址置信度",
               "总金额", "金额置信度",
               "处理耗时(ms)", "创建时间", "更新时间"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    status_map = {"processed": "已处理", "approved": "已通过",
                  "rejected": "已拒绝", "pending": "待处理"}

    for row_idx, r in enumerate(rows, 2):
        sample_id = _extract_sample_id(r["image_path"])
        fields = all_fields.get(r["id"], {})

        values = [
            sample_id,
            status_map.get(r["status"], r["status"]),
            "异常" if r["risk_level"] == "anomaly" else "正常",
        ]
        for fk in ["company", "date", "address", "total"]:
            f = fields.get(fk)
            values.append(f["final_value"] if f else "")
            values.append(round(f["confidence"] * 100, 1) if f and f["confidence"] else "")

        values.extend([
            round(r["processing_time_ms"], 1) if r["processing_time_ms"] else "",
            r["created_at"] or "",
            r["updated_at"] or "",
        ])

        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border

    # 设置列宽
    widths = [12, 8, 8, 30, 10, 15, 10, 40, 10, 12, 10, 12, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"invoices_{status or 'all'}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.on_event("startup")
async def startup():
    init_db()
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
